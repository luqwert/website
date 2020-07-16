#!/usr/bin/env python
# _*_ coding:utf-8 _*_
# @Author  : lusheng
import time

from flask import Flask,render_template,request,url_for, Response,redirect
import platform
from datetime import timedelta
import os
from pylab import *
import matplotlib as mpl
from xlrd import open_workbook
from xlrd import xldate_as_datetime
import datetime
from openpyxl import load_workbook
import json

def xlstolist(s):
    datelist = []
    for row in range(s.nrows):
        if type(s.cell_value(row, 0)) is str:
            date = s.cell_value(row, 0)
        else:
            date = xldate_as_datetime(s.cell_value(row, 0), 0)
            date = str(date.strftime('%Y-%m-%d'))
        datelist.append(date)
    return datelist

def creation_date(path_to_file):
    """
    Try to get the date that a file was created, falling back to when it was
    last modified if that isn't possible.
    See http://stackoverflow.com/a/39501288/1709587 for explanation.
    """
    if platform.system() == 'Windows':
        timestamp = os.path.getmtime(path_to_file)
        timeStruct = time.localtime(timestamp)
        strftime = time.strftime('%Y-%m-%d', timeStruct)
        return strftime
    else:
        stat = os.stat(path_to_file)
        try:
            timestamp = stat.st_birthtime
            timeStruct = time.localtime(timestamp)
            strftime = time.strftime('%Y-%m-%d', timeStruct)
            return strftime
        except AttributeError:
            # We're probably on Linux. No easy way to get creation dates here,
            # so we'll settle for when its content was last modified.
            timestamp = stat.st_mtime
            timeStruct = time.localtime(timestamp)
            strftime = time.strftime('%Y-%m-%d', timeStruct)
            return strftime




app = Flask(__name__)
UPLOAD_FOLDER = 'static/周报材料'
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER  # 设置文件上传的目标文件夹
basedir = os.path.abspath(os.path.dirname(__file__))  # 获取当前项目的绝对路径
ALLOWED_EXTENSIONS = set(['txt', 'png', 'jpg', 'xls', 'JPG', 'PNG', 'xlsx', 'gif', 'GIF', 'ppt', 'docx', 'mp4', 'flv', 'pdf', 'doc','docx','jpeg','JPEG'])  # 允许上传的文件后缀
app.secret_key = 'wqewqrwqeqwq'

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1] in ALLOWED_EXTENSIONS


@app.route('/')
def index():
    # mdate = creation_date('E:/python/website/static/各产品周分析报告.pdf')
    return render_template('index2.html')

@app.route('/report/')
def report():
    # ptime = creation_date('E:\\python\\website\\static\\周报材料\\华诚金属.txt')
    ptime = creation_date('./static/周报材料/华诚金属.txt')
    # excel_path = 'E:\\python\\website\\static\\周报材料\\周分析会议数据.xlsx'
    excel_path = './static/周报材料/周分析会议数据.xlsx'
    wb = load_workbook(excel_path)
    ws = wb.get_sheet_by_name("普氏、MYSTEEL指数")
    print(ws['A%d' % ws.max_row].value)
    date1 = str(ws['A%d' % ws.max_row].value)[:10]
    date1_62 = str(ws['B%d' % ws.max_row].value)
    date1_58 = str(ws['C%d' % ws.max_row].value)
    date2 = str(ws['A%d' % (ws.max_row - 5)].value)[:10]
    date2_62 = str(ws['B%d' % (ws.max_row - 5)].value)
    date2_58 = str(ws['C%d' % (ws.max_row - 5)].value)
    wb.close()
    # print(date1,date2,date1_58,date1_62, date2_58,date2_62)
    if float(date1_62) - float(date2_62) >= 0:
        updown_62 = '上涨'
        diff_62 = round(float(date1_62) - float(date2_62),2)
    else:
        updown_62 = '下跌'
        diff_62 = round(-(float(date1_62) - float(date2_62)),2)

    if float(date1_58) - float(date2_58) >= 0:
        updown_58 = '上涨'
        diff_58 = round(float(date1_58) - float(date2_58),2)
    else:
        updown_58 = '下跌'
        diff_58 = round(-(float(date1_58) - float(date2_58)),2)

    mysteeltext = []
    # with open('E:\\python\\website\\static\\周报材料\\mysteel.txt', 'r', encoding='utf-8') as f_mysteel:
    with open('./static/周报材料/mysteel.txt', 'r', encoding='utf-8') as f_mysteel:

        for line in f_mysteel:
            mysteeltext.append(line.strip('\n').split(','))
        # print(mysteeltext)
        kucun_text = mysteeltext[0][0]
        # print(kucun_text)
        kaigong_text = mysteeltext[2][0]
        # print(kaigong_text)
        fenxi_text = mysteeltext[4][0] + mysteeltext[6][0]
        # print(fenxi_text)
        haiyunfei_text = mysteeltext[10][0]
        # print(haiyunfei_text)
        feigang_text = mysteeltext[8][0]
        # print(feigang_text)

    # with open('E:\\python\\website\\static\\周报材料\\锰矿.txt', 'r', encoding='utf-8') as f_mengkuang:
    with open('./static/周报材料/锰矿.txt', 'r', encoding='utf-8') as f_mengkuang:
        mengkuang_text = f_mengkuang.read()
    # with open('E:\\python\\website\\static\\周报材料\\硅锰.txt', 'r', encoding='utf-8') as f_guimeng:
    with open('./static/周报材料/硅锰.txt', 'r', encoding='utf-8') as f_guimeng:
        guimeng_text = f_guimeng.read()

    stock2 = []
    line = []
    # wb = load_workbook('E:\\python\\website\\static\\周报材料\\cnfeol1.xlsx')
    wb = load_workbook('./static/周报材料/cnfeol1.xlsx')
    ws = wb.active
    for row in range(ws.max_row):
        for col in range(ws.max_column):
            line.append(ws.cell(row=(row + 1), column=(col + 1)).value)
        stock2.append(line)
        line = []
    wb.close()
    # print(stock2)
    menggkuang_kucun = stock2[-1][-2]
    # print(menggkuang_kucun)
    if float(stock2[-1][-2]) - float(stock2[-1][-3]) > 0:
        updown_mengkuang = '增加'
        diff_mengkuang = str(stock2[-1][-1]) + '吨'
    elif float(stock2[-1][-2]) - float(stock2[-1][-3]) < 0:
        updown_mengkuang = '减少'
        diff_mengkuang = str(stock2[-1][-1])[1:] + '吨'
    elif float(stock2[-1][-2]) - float(stock2[-1][-3]) == 0:
        updown_mengkuang = '不变'
        diff_mengkuang = ''

    mengkuang_price = []
    line3 = []
    # wb = load_workbook('E:\\python\\website\\static\\周报材料\\cnfeol2.xlsx')
    wb = load_workbook('./static/周报材料/cnfeol2.xlsx')
    ws = wb.active
    for row in range(ws.max_row):
        for col in range(ws.max_column):
            line3.append(ws.cell(row=(row + 1), column=(col + 1)).value)
        mengkuang_price.append(line3)
        line3 = []
    wb.close()
    # print(mengkuang_price)

    guimeng_price = []
    line4 = []
    # wb = load_workbook('E:\\python\\website\\static\\周报材料\\cnfeol3.xlsx')
    wb = load_workbook('./static/周报材料/cnfeol3.xlsx')
    ws = wb.active
    for row in range(ws.max_row):
        for col in range(ws.max_column):
            line4.append(ws.cell(row=(row + 1), column=(col + 1)).value)
        guimeng_price.append(line4)
        line4 = []
    wb.close()
    # print(guimeng_price)

    # with open('E:\\python\\website\\static\\周报材料\\华诚金属.txt', 'r', encoding='utf-8') as f_mengpian:
    with open('./static/周报材料/华诚金属.txt', 'r', encoding='utf-8') as f_mengpian:
        mengpian_text = f_mengpian.read()

    return render_template('report.html',
                           ptime=ptime,
                           date1=date1,
                           date2=date2,
                           date1_62=date1_62,
                           date2_62=date2_62,
                           date1_58=date1_58,
                           date2_58=date2_58,
                           diff_58=diff_58,
                           diff_62=diff_62,
                           updown_58=updown_58,
                           updown_62=updown_62,
                           kucun_text=kucun_text,
                           kaigong_text=kaigong_text,
                           fenxi_text=fenxi_text,
                           haiyunfei_text=haiyunfei_text,
                           feigang_text=feigang_text,
                           mengkuang_text=mengkuang_text,
                           guimeng_text=guimeng_text,
                           stock2=stock2,
                           menggkuang_kucun=menggkuang_kucun,
                           mengkuang_price=mengkuang_price,
                           updown_mengkuang=updown_mengkuang,
                           diff_mengkuang=diff_mengkuang,
                           guimeng_price=guimeng_price,
                           mengpian_text=mengpian_text,
                           )

@app.route('/shujuchaxun/')
def shujuchaxun():

    return render_template('shujuchaxun.html')


@app.route('/platts',methods=['POST'])
def platts():
    # 获取数据
    sdate = request.form.get('sdate')
    edate = request.form.get('edate')

    x_data = []
    x_data_str = []
    y_data1 = []
    y_data2 = []
    price_list = [['日期', '62指数', '58指数']]
    mpl.rcParams['font.sans-serif'] = ['SimHei']
    mpl.rcParams['axes.unicode_minus'] = False
    # wb = open_workbook('E:\\python\\website\\static\\周报材料\\周分析会议数据.xlsx')
    wb = open_workbook('./static/周报材料/周分析会议数据.xlsx')
    s = wb.sheet_by_name(u'普氏、MYSTEEL指数')

    def xlstolist(s):
        datelist = []
        for row in range(s.nrows):
            if type(s.cell_value(row, 0)) is str:
                date = s.cell_value(row, 0)
            else:
                date = xldate_as_datetime(s.cell_value(row, 0), 0)
                date = str(date.strftime('%Y-%m-%d'))
            datelist.append(date)
        return datelist

    datelist = xlstolist(s)
    if sdate in datelist:
        sdatea = sdate
    else:
        # print(datetime.datetime.strptime(sdate, "%Y-%m-%d"))
        # print(datetime.datetime.strptime(datelist[-1], "%Y-%m-%d"))
        if datetime.datetime.strptime(sdate, "%Y-%m-%d") >= datetime.datetime.strptime(datelist[-1], "%Y-%m-%d"):
            sdatea = datetime.datetime.strptime(datelist[-1], "%Y-%m-%d")
        else:
            for i in range(len(datelist)):
                sdate = datetime.datetime.strptime(str(sdate), "%Y-%m-%d") - datetime.timedelta(days=-1)
                sdate = str(sdate)[:10]
                if sdate in datelist:
                    # print(sdate)
                    sdatea = sdate
                    break

    for n in range(len(datelist)):
        if sdatea == datelist[n]:
            start_row = n
            # print(start_row)
            break

    if edate in datelist:
        edatea = edate
        print(edatea)
    else:
        print(datetime.datetime.strptime(edate, "%Y-%m-%d"))
        print(datetime.datetime.strptime(datelist[1], "%Y-%m-%d"))
        if datetime.datetime.strptime(edate, "%Y-%m-%d") <= datetime.datetime.strptime(datelist[1], "%Y-%m-%d"):
            edatea = datetime.datetime.strptime(datelist[-1], "%Y-%m-%d")
        else:
            for i in range(len(datelist)):
                edate = datetime.datetime.strptime(str(edate), "%Y-%m-%d") - datetime.timedelta(days=1)
                edate = str(edate)[:10]
                if edate in datelist:
                    print(edate)
                    edatea = edate
                    break

    for n in range(len(datelist)):
        if edatea == datelist[n]:
            end_row = n
            print(end_row)
            break

    print(start_row, end_row)
    for row in range(start_row, end_row+1):
        if type(s.cell_value(row, 0)) is str:
            date = s.cell_value(row, 0)
        else:
            date = xldate_as_datetime(s.cell_value(row, 0), 0)
            date = str(date.strftime('%Y-%m-%d'))
        x_data_str.append(date)
        y_data1.append(s.cell_value(row, 1))
        y_data2.append(s.cell_value(row, 2))
        x_data = [datetime.datetime.strptime(str(d), '%Y-%m-%d').date() for d in x_data_str]
    for n in reversed(range(len(x_data_str))):
        price_list.append([x_data_str[n], y_data1[n], y_data2[n]])

    print(x_data_str)
    print(x_data)
    print(len(y_data1))
    print(len(y_data2))
    print(price_list)

    # 转换成JSON数据格式
    jsonData = {}
    xdays = []
    yvalues1 = []
    yvalues2 = []
    for data in x_data_str:
        # xdays.append(str(data[0]))
        xdays.append(data)
    for ydata1 in y_data1:
        yvalues1.append(ydata1)
    for ydata2 in y_data2:
        yvalues2.append(ydata2)
    print(xdays)
    print(yvalues1)
    print(yvalues2)

    jsonData['xdays'] = xdays
    jsonData['yvalues1'] = yvalues1
    jsonData['yvalues2'] = yvalues2
    jsonData['price_list'] = price_list
    # json.dumps()用于将dict类型的数据转成str，因为如果直接将dict类型的数据写入json会发生报错，因此将数据写入时需要用到该函数。
    j = json.dumps(jsonData)

    # 在浏览器上渲染my_template.html模板（为了查看输出的数据）
    return(j)

@app.route('/mengpian',methods=['POST'])
def mengpian():
    # 获取数据
    sdate = request.form.get('sdate')
    edate = request.form.get('edate')

    x_data = []
    x_data_str = []
    y_data1 = []
    y_data2 = []
    price_list = [['日期', '价格', '平均价格']]
    # mpl.rcParams['font.sans-serif'] = ['SimHei']
    # mpl.rcParams['axes.unicode_minus'] = False
    # wb = open_workbook('E:\\python\\website\\static\\周报材料\\周分析会议数据.xlsx')
    wb = open_workbook('./static/周报材料/周分析会议数据.xlsx')
    s = wb.sheet_by_name(u'电解锰片价格')

    def xlstolist(s):
        datelist = []
        for row in range(s.nrows):
            if type(s.cell_value(row, 0)) is str:
                date = s.cell_value(row, 0)
            else:
                date = xldate_as_datetime(s.cell_value(row, 0), 0)
                date = str(date.strftime('%Y-%m-%d'))
            datelist.append(date)
        return datelist

    datelist = xlstolist(s)
    if sdate in datelist:
        sdatea = sdate
    else:
        # print(datetime.datetime.strptime(sdate, "%Y-%m-%d"))
        # print(datetime.datetime.strptime(datelist[-1], "%Y-%m-%d"))
        if datetime.datetime.strptime(sdate, "%Y-%m-%d") >= datetime.datetime.strptime(datelist[-1], "%Y-%m-%d"):
            sdatea = datetime.datetime.strptime(datelist[-1], "%Y-%m-%d")
        else:
            for i in range(len(datelist)):
                sdate = datetime.datetime.strptime(str(sdate), "%Y-%m-%d") - datetime.timedelta(days=-1)
                sdate = str(sdate)[:10]
                if sdate in datelist:
                    # print(sdate)
                    sdatea = sdate
                    break

    for n in range(len(datelist)):
        if sdatea == datelist[n]:
            start_row = n
            # print(start_row)
            break

    if edate in datelist:
        edatea = edate
        print(edatea)
    else:
        print(datetime.datetime.strptime(edate, "%Y-%m-%d"))
        print(datetime.datetime.strptime(datelist[1], "%Y-%m-%d"))
        if datetime.datetime.strptime(edate, "%Y-%m-%d") <= datetime.datetime.strptime(datelist[1], "%Y-%m-%d"):
            edatea = datetime.datetime.strptime(datelist[-1], "%Y-%m-%d")
        else:
            for i in range(len(datelist)):
                edate = datetime.datetime.strptime(str(edate), "%Y-%m-%d") - datetime.timedelta(days=1)
                edate = str(edate)[:10]
                if edate in datelist:
                    print(edate)
                    edatea = edate
                    break

    for n in range(len(datelist)):
        if edatea == datelist[n]:
            end_row = n
            print(end_row)
            break

    print(start_row, end_row)
    for row in range(start_row, end_row+1):
        if type(s.cell_value(row, 0)) is str:
            date = s.cell_value(row, 0)
        else:
            date = xldate_as_datetime(s.cell_value(row, 0), 0)
            date = str(date.strftime('%Y-%m-%d'))
        x_data_str.append(date)
        y_data1.append(s.cell_value(row, 1))
        y_data2.append(s.cell_value(row, 2))
        x_data = [datetime.datetime.strptime(str(d), '%Y-%m-%d').date() for d in x_data_str]
    for n in reversed(range(len(x_data_str))):
        price_list.append([x_data_str[n], y_data1[n], y_data2[n]])

    # print(x_data_str)
    # print(x_data)
    # print(len(y_data1))
    # print(len(y_data2))
    # print(price_list)

    # 转换成JSON数据格式
    jsonData = {}
    xdays = []
    yvalues1 = []
    yvalues2 = []
    for data in x_data_str:
        # xdays.append(str(data[0]))
        xdays.append(data)
    for ydata1 in y_data1:
        yvalues1.append(ydata1)
    for ydata2 in y_data2:
        yvalues2.append(ydata2)
    print(xdays)
    print(yvalues1)
    print(yvalues2)

    jsonData['xdays'] = xdays
    jsonData['yvalues1'] = yvalues1
    jsonData['yvalues2'] = yvalues2
    jsonData['price_list'] = price_list
    # json.dumps()用于将dict类型的数据转成str，因为如果直接将dict类型的数据写入json会发生报错，因此将数据写入时需要用到该函数。
    j = json.dumps(jsonData)

    # 在浏览器上渲染my_template.html模板（为了查看输出的数据）
    return(j)


@app.route('/admin/', methods=['POST', 'GET'], strict_slashes=False)
def admin():
    if request.method == 'POST':
        file_dir = os.path.join(basedir, app.config['UPLOAD_FOLDER'])  # 拼接成合法文件夹地址
        if not os.path.exists(file_dir):
            os.makedirs(file_dir)  # 文件夹不存在就创建
        files = request.files.getlist('myfile')# 从表单的file字段获取文件，myfile为该表单的name值
        for f in files:
            if f and allowed_file(f.filename):  # 判断是否是允许上传的文件类型
                fname = f.filename
                print(os.path.join(file_dir, fname))
                f.save(os.path.join(file_dir, fname))  # 保存文件到upload目录
            else:
                pass
        return render_template('admin.html', status='OK')

    else:
        return render_template('admin.html')


@app.route('/ciyun/')
def ciyun():
    return render_template('ciyun.html')

@app.route('/kuaidi/')
def kuaidi():
    return render_template('kuaidi.html')

@app.route('/yinzhang/')
def yinzhang():
    return render_template('yinzhang.html')

@app.route('/wenbentiqu/')
def wenbentiqu():
    return render_template('wenbentiqu.html')

if __name__ == '__main__':
    # app.run(host='0.0.0.0',port='80')
    app.run(debug='ture')
